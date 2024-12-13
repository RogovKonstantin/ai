import pandas as pd
import pickle
import openpyxl
from openpyxl.styles import PatternFill, Alignment


def load_dataset(file_path):
    with open(file_path, 'r') as file:
        lines = file.readlines()
    headers = lines[0].strip().split(',')
    data = [list(map(float, line.strip().split(','))) for line in lines[1:]]
    return headers, data


headers, data = load_dataset('../../datasets/normalized_shuffled_test.csv')

feature_indices = [i for i in range(len(headers)) if headers[i] != 'TenYearCHD']
target_index = headers.index('TenYearCHD')
X = [[row[i] for i in feature_indices] for row in data]
y = [row[target_index] for row in data]

with open('../model_lib_weights.pkl', 'rb') as f:
    model_lib = pickle.load(f)


def load_weights(file_path):
    with open(file_path, 'r') as file:
        weights = list(map(float, file.read().strip().split(',')))
    return weights


weights = load_weights('../model_manual_weights.txt')


def sigmoid(z):
    return 1 / (1 + (2.71828 ** -z))


def predict_manual(X, weights):
    X = [[1.0] + row for row in X]
    predictions = []
    for row in X:
        z = sum(w * x for w, x in zip(weights, row))
        predictions.append(1 if sigmoid(z) >= 0.5 else 0)
    return predictions


predictions_lib = model_lib.predict(pd.DataFrame(X, columns=headers[:-1]))
predictions_manual = predict_manual(X, weights)

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "Model Comparison"

color_both_right = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
color_lib_right = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
color_manual_right = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
color_none_right = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")

accuracy_lib = sum([1 for i in range(len(y)) if predictions_lib[i] == y[i]]) / len(y)
accuracy_manual = sum([1 for i in range(len(y)) if predictions_manual[i] == y[i]]) / len(y)
cases_both_correct = sum(1 for i in range(len(y)) if predictions_lib[i] == y[i] and predictions_manual[i] == y[i])
cases_lib_only = sum(1 for i in range(len(y)) if predictions_lib[i] == y[i] and predictions_manual[i] != y[i])
cases_manual_only = sum(1 for i in range(len(y)) if predictions_manual[i] == y[i] and predictions_lib[i] != y[i])
cases_neither = sum(1 for i in range(len(y)) if predictions_lib[i] != y[i] and predictions_manual[i] != y[i])

sheet.append(["Точность (с библиотеками)", accuracy_lib])
sheet.append(["Точность (без библиотек)", accuracy_manual])
sheet.append(["обе модели предсказали верно", cases_both_correct / len(y)])
sheet.append(["Предсказала верно модель с библиотеками", cases_lib_only / len(y)])
sheet.append(["Предсказала верно модель без библиотек", cases_manual_only / len(y)])
sheet.append(["Обе модели предсказали неверно", cases_neither / len(y)])

sheet.append([])

sheet["A3"].fill = color_both_right
sheet["A4"].fill = color_lib_right
sheet["A5"].fill = color_manual_right
sheet["A6"].fill = color_none_right

sheet.append(headers + ["", "С библиотеками", "Без библиотек"])

for cell in sheet[sheet.max_row]:
    cell.alignment = Alignment(horizontal="center")

start_row = sheet.max_row + 1

for i, row in enumerate(X):
    actual = int(y[i])
    pred_lib = int(predictions_lib[i])
    pred_manual = int(predictions_manual[i])

    if pred_lib == actual and pred_manual == actual:
        fill = color_both_right
    elif pred_lib == actual and pred_manual != actual:
        fill = color_lib_right
    elif pred_manual == actual and pred_lib != actual:
        fill = color_manual_right
    else:
        fill = color_none_right

    current_row = start_row + i
    sheet.append(row + [actual, "", pred_lib, pred_manual])

    for col in range(1, len(row) + 5):
        sheet.cell(row=current_row, column=col).fill = fill

for col in sheet.columns:
    values = [str(cell.value) for cell in col if cell.value is not None]
    if values:
        max_length = max(len(value) for value in values)
        adjusted_width = max_length + 2
        sheet.column_dimensions[col[0].column_letter].width = adjusted_width

wb.save("../model_comparison.xlsx")
